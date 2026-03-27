"""
Laiteluettelo - Excel-tiedoston generointi
Tuottaa tyylitellyn .xlsx-tiedoston joka on valmis kopioitavaksi viralliseen pohjaan.
"""
from __future__ import annotations
from datetime import date
from pathlib import Path
from typing import Optional, Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .equipment_schema import ExtractedUnit, ExtractedComponent, load_equipment_config


# ── Väripaletti ──────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":   "1F3864",   # Tummansininen otsikko
    "header_fg":   "FFFFFF",
    "subheader_bg":"2E75B6",   # Sininen alaotsikko
    "subheader_fg":"FFFFFF",
    "col_header_bg":"D6E4F0",  # Sarakeotsikoiden tausta
    "col_header_fg":"1F3864",

    # Laitetyyppikohtaiset värit
    "SP": "E2EFDA",   # Vaaleanvihreä – sulkupelti
    "FG": "F2F2F2",   # Harmaa – peltimoottori
    "SU": "FFF2CC",   # Keltainen – suodatin
    "LTO_supply":  "DAEEF3",  # Vaaleansininen – LTO tulopuoli
    "LTO_exhaust": "E9D7F5",  # Violetti – LTO poistopuoli
    "TF": "FCE4D6",   # Oranssi – puhallin
    "LP": "FFE6E6",   # Pinkki – lämmityspatteri
    "JP": "D9F2FF",   # Syaani – jäähdytyspatteri
    "AV": "EAD1DC",   # Roosa – äänenvaimennin
    "HPE": "E8D4F8",  # Vaaleanvioletti – esilämmityspatteri
    "HPO": "D5E8F7",  # Vaaleansininen – jälkilämmityspatteri
    "SOUND": "F0E8D8", # Vaalea beige – äänitiedot
    "default": "FFFFFF",
}

# Ohuet reunaviivat
THIN = Side(style="thin", color="BFBFBF")
THICK = Side(style="medium", color="1F3864")


def make_border(left=THIN, right=THIN, top=THIN, bottom=THIN):
    return Border(left=left, right=right, top=top, bottom=bottom)


OUTER_BORDER = make_border(THICK, THICK, THICK, THICK)
INNER_BORDER = make_border()


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="000000", name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Sarakkeiden määritys ──────────────────────────────────────────────────────
# Jokainen tuple: (otsikko, leveys, arvo-kenttä tai None)
COLUMNS = [
    # (otsikko, sarakeleveys)
    ("TUNNUS",            10),
    ("LAITE",             20),
    ("PUOLI",             10),
    ("ILMAVIRTA\n(m³/s)", 10),
    ("ILMA ΔP\n(Pa)",     10),
    ("ALKU ΔP\n(Pa)",      9),
    ("LOPPU ΔP\n(Pa)",     9),
    ("SUOD.\nLUOKKA",     12),
    ("ILMA ENS\n(°C)",    10),
    ("ILMA JÄLK\n(°C)",   10),
    ("NESTE\nVIRTA (l/s)",11),
    ("NESTE ΔP\n(kPa)",   10),
    ("NESTE\nMENO (°C)",  10),
    ("NESTE\nPALUU (°C)", 10),
    ("TEHO\n(kW)",         9),
    ("JÄNNITE/\nVIRTA",   13),
    ("HYÖTS.\n(%)",        9),
    ("HUOMIOT",           28),
]

# Kartoitus: sarakeindeksi → (laitetyyppi tai "*", data-avain)
# Täyttää arvon automaattisesti oikeaan sarakkeeseen
FIELD_MAP: dict[int, list[tuple[str, str]]] = {
    3:  [("TF", "ilmamaara")],
    4:  [("SP", "ilma_dp"), ("SU", "ilma_dp_mitoitus"), ("TF", "mitoituspaine"),
         ("LP", "ilma_dp"), ("JP", "ilma_dp"), ("AV", "ilma_dp"),
         ("HPE", "ilma_dp"), ("HPO", "ilma_dp"),
         ("LTO_supply", "ilma_dp"), ("LTO_exhaust", "ilma_dp")],
    5:  [("SU", "ilma_dp_alku")],
    6:  [("SU", "ilma_dp_loppu")],
    7:  [("SU", "suodatinluokka")],
    8:  [("LP", "ilma_lampotila_ennen"), ("JP", "ilma_lampotila_ennen"),
         ("HPE", "ilma_lampotila_ennen"), ("HPO", "ilma_lampotila_ennen"),
         ("LTO_supply", "ilma_lampotila_ennen"), ("LTO_exhaust", "ilma_lampotila_ennen")],
    9:  [("LP", "ilma_lampotila_jalkeen"), ("JP", "ilma_lampotila_jalkeen"),
         ("HPE", "ilma_lampotila_jalkeen"), ("HPO", "ilma_lampotila_jalkeen"),
         ("LTO_supply", "ilma_lampotila_jalkeen"), ("LTO_exhaust", "ilma_lampotila_jalkeen")],
    10: [("LP", "nestevirta"), ("JP", "nestevirta"),
         ("HPE", "nestevirta"), ("HPO", "nestevirta")],
    11: [("LP", "neste_dp"), ("JP", "neste_dp"),
         ("HPE", "neste_dp"), ("HPO", "neste_dp")],
    12: [("LP", "neste_meno"), ("JP", "neste_meno"),
         ("HPE", "neste_meno"), ("HPO", "neste_meno")],
    13: [("LP", "neste_paluu"), ("JP", "neste_paluu"),
         ("HPE", "neste_paluu"), ("HPO", "neste_paluu")],
    14: [("TF", "sahkoteho")],
    15: [("TF", "jannite_virta")],
    16: [("LTO_supply", "hyotysuhde_en308")],
}


def _get_row_type_key(comp: ExtractedComponent, row_idx: int) -> str:
    """Palauta väriavaimen tyyppi rivi-indeksin mukaan."""
    if comp.type_prefix == "LTO":
        return "LTO_supply" if row_idx == 0 else "LTO_exhaust"
    return comp.type_prefix


def _get_cell_value(data: dict, type_key: str, col_idx: int) -> Any:
    """Hae arvo datasta sarakeindeksin ja tyypin perusteella."""
    candidates = FIELD_MAP.get(col_idx, [])
    for tk, field in candidates:
        if tk == type_key and field in data:
            val = data[field]
            if val is None:
                return None
            # Numeerinen formatointi
            if isinstance(val, float):
                return round(val, 2)
            return val
    return None


def _format_additional_data(data: dict, type_key: str) -> str:
    """
    Muodosta lisätiedot (kosteus, äänitiedot) HUOMIOT-sarakkeelle.
    Sisältää tiedot jotka eivät mahdu muihin sarakkeisiin.
    """
    notes = []

    # Kosteus (HPE, HPO, LTO, LP, JP)
    if "ilma_kosteus_ennen" in data and data["ilma_kosteus_ennen"] is not None:
        before = data["ilma_kosteus_ennen"]
        after = data.get("ilma_kosteus_jalkeen")
        if after is not None:
            notes.append(f"Kosteus: {before}% → {after}%")
        else:
            notes.append(f"Kosteus ennen: {before}%")
    elif "ilma_kosteus_jalkeen" in data and data["ilma_kosteus_jalkeen"] is not None:
        notes.append(f"Kosteus jälkeen: {data['ilma_kosteus_jalkeen']}%")

    # Äänitiedot (SOUND)
    if "aani_data_json" in data and data["aani_data_json"]:
        try:
            audio_data = data["aani_data_json"]
            if isinstance(audio_data, dict):
                locations = list(audio_data.keys())
                notes.append(f"Äänitiedot ({len(locations)} mittauspistettä)")
                # Näytä kokonais dB-A arvot
                for loc in sorted(locations)[:3]:  # Näytä max 3 mittauspistettä
                    if isinstance(audio_data[loc], dict) and "kokonais_dB_A" in audio_data[loc]:
                        db_a = audio_data[loc]["kokonais_dB_A"]
                        loc_pretty = loc.replace("_", " ")
                        notes.append(f"  {loc_pretty}: {db_a} dB(A)")
        except (TypeError, KeyError):
            notes.append("Äänitiedot saatavilla")

    return "\n".join(notes)


def _write_row(
    ws: Worksheet,
    row: int,
    comp: ExtractedComponent,
    component_row_idx: int,
    config: dict,
) -> None:
    """Kirjoita yksi laiterivi Exceliin."""
    comp_row = comp.rows[component_row_idx]
    data = comp_row.data
    type_key = _get_row_type_key(comp, component_row_idx)
    color = COLORS.get(type_key, COLORS["default"])
    fill = _fill(color)

    # Sarake 1: TUNNUS
    c = ws.cell(row=row, column=1, value=comp.code)
    c.fill = fill; c.font = _font(bold=True, size=10)
    c.alignment = _align("center"); c.border = INNER_BORDER

    # Sarake 2: LAITE
    name = comp.name
    if comp_row.row_label:
        name = f"{comp.name} / {comp_row.row_label}"
    c = ws.cell(row=row, column=2, value=name)
    c.fill = fill; c.font = _font(size=10)
    c.alignment = _align("left"); c.border = INNER_BORDER

    # Sarake 3: PUOLI
    side_fi = {"supply": "Tulo", "exhaust": "Poisto", "both": "Tulo+Poisto"}.get(comp.side, comp.side)
    if comp.type_prefix == "LTO":
        side_fi = "Tulo" if component_row_idx == 0 else "Poisto"
    c = ws.cell(row=row, column=3, value=side_fi)
    c.fill = fill; c.font = _font(size=9)
    c.alignment = _align("center"); c.border = INNER_BORDER

    # Sarakkeet 4–17: tekniset tiedot
    for col_idx in range(3, len(COLUMNS)):
        # HUOMIOT-sarake (viimeinen) saa lisätiedot (kosteus, äänitiedot)
        if col_idx == len(COLUMNS) - 1:
            val = _format_additional_data(data, type_key)
        else:
            val = _get_cell_value(data, type_key, col_idx)

        c = ws.cell(row=row, column=col_idx + 1, value=val)
        c.fill = fill
        c.font = _font(size=10)
        c.alignment = _align("left" if col_idx == len(COLUMNS) - 1 else "center", wrap=True)
        c.border = INNER_BORDER
        # Tyhjä solu haalean harmaana
        if not val:
            c.fill = _fill("F7F7F7")


def generate_excel(
    unit: ExtractedUnit,
    output_path: Optional[str | Path] = None,
) -> Path:
    """
    Luo tyylitelty laiteluettelo-Excel IV-koneelle.
    Palauttaa tallennetun tiedostopolun.
    """
    config = load_equipment_config()
    wb = Workbook()
    ws = wb.active
    # Sanitize sheet title (Excel doesn't allow: ? * [ ] / \)
    sanitized_code = unit.unit_code.replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace("/", "").replace("\\", "") or "TK"
    ws.title = sanitized_code[:31]  # Excel sheet name max 31 chars

    # Lukitse ylimmät rivit jäätymistä varten
    ws.freeze_panes = "A4"

    # ── RIVI 1: Projektiotsikko ───────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    header_text = f"LAITELUETTELO – {unit.unit_code}"
    if unit.project:
        header_text += f"  |  {unit.project}"
    if unit.manufacturer and unit.model:
        header_text += f"  |  {unit.manufacturer} {unit.model}"
    c = ws.cell(row=1, column=1, value=header_text)
    c.fill = _fill(COLORS["header_bg"])
    c.font = _font(bold=True, size=13, color=COLORS["header_fg"])
    c.alignment = _align("center")
    c.border = make_border(THICK, THICK, THICK, THICK)
    ws.row_dimensions[1].height = 22

    # ── RIVI 2: Alaotsikko (pvm + ilmavirrat) ────────────────────────────────
    ws.merge_cells("A2:R2")
    sub_parts = [f"Päivitetty: {date.today().strftime('%d.%m.%Y')}"]
    if unit.raw_supply_airflow:
        sub_parts.append(f"Tuloilma: {unit.raw_supply_airflow} m³/s")
    if unit.raw_exhaust_airflow:
        sub_parts.append(f"Poistoilma: {unit.raw_exhaust_airflow} m³/s")
    if unit.source_pdf:
        sub_parts.append(f"Lähde: {Path(unit.source_pdf).name}")
    c = ws.cell(row=2, column=1, value="   ".join(sub_parts))
    c.fill = _fill(COLORS["subheader_bg"])
    c.font = _font(bold=False, size=9, color=COLORS["subheader_fg"])
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 16

    # ── RIVI 3: Sarakkeotsikot ────────────────────────────────────────────────
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=3, column=col_idx, value=label)
        c.fill = _fill(COLORS["col_header_bg"])
        c.font = _font(bold=True, size=9, color=COLORS["col_header_fg"])
        c.alignment = _align("center", wrap=True)
        c.border = make_border(THICK, THICK, THICK, THICK)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 34

    # ── RIVIT 4+: Laitteet ────────────────────────────────────────────────────
    current_row = 4
    for comp in unit.components:
        for row_idx, _ in enumerate(comp.rows):
            _write_row(ws, current_row, comp, row_idx, config)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # Ohut viiva laitteiden välillä
        # (toteutettu jo reunoissa)

    # ── Tulostusalue ─────────────────────────────────────────────────────────
    ws.print_area = f"A1:{get_column_letter(len(COLUMNS))}{current_row - 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # ── Tallenna ──────────────────────────────────────────────────────────────
    if output_path is None:
        output_dir = Path(__file__).parent.parent / "output"
        output_dir.mkdir(exist_ok=True)
        # Sanitize filename (remove invalid chars: ? * [ ] / \ : ; < > |)
        safe_code = unit.unit_code.replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace("/", "").replace("\\", "").replace(":", "").replace(";", "").replace("<", "").replace(">", "").replace("|", "") or "TK"
        output_path = output_dir / f"laiteluettelo_{safe_code}.xlsx"

    output_path = Path(output_path)
    wb.save(output_path)
    return output_path
